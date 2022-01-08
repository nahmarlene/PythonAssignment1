import openpyxl as xl 
import csv
wb = xl.load_workbook('employeedata.xlsx')
sheet=wb['Sheet1']
old_email = 'helpinghands.cm'
new_email = 'handsinhands.org'
for i in range (2, sheet.max_row+1):
    cell = sheet.cell(i,3)
    if old_email in cell.value:
       updated_Email=(cell.value).replace(old_email, new_email)

       sheet.cell(i,3).value = updated_Email

wb.save('employee.csv')
wb.save('employee.xlsx')